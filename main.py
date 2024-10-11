import asyncio
import os
import random
import string
import sys
import threading

import discord
import openpyxl
from discord.ext import commands
from filelock import FileLock
from flask import Flask, jsonify, request
from flask_cors import CORS

app = Flask(__name__)

CORS(app)

API_KEY = os.getenv("API_KEY", "TapSs@14023010.com")
TOKEN = os.getenv(
    "DISCORD_TOKEN",
    "MTI5NDAxNjE4MDE5NzEzMDI1MQ.GB93Tg.wTqieZt3A3BvSMxP6cNr5O7LVsNXRzslBaVGQM",
)
intents = discord.Intents.default()
intents.message_content = True
intents.guilds = True
intents.members = True
bot = commands.Bot(command_prefix="!", intents=intents)

ROLE_NAME = "Subscriber"
TOKEN_FILE = "token.xlsx"
USER_FILE = "users.xlsx"

token_cache = set()


def generate_unique_token(length=10):
    return "".join(random.choices(string.ascii_letters + string.digits, k=length))


def load_existing_tokens():
    existing_tokens = set()

    if os.path.exists(TOKEN_FILE):
        workbook = openpyxl.load_workbook(TOKEN_FILE)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=1, max_col=1):
            token = row[0].value
            if token:
                existing_tokens.add(token)

    if os.path.exists(USER_FILE):
        workbook = openpyxl.load_workbook(USER_FILE)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, max_col=2):
            token = row[1].value
            if token:
                existing_tokens.add(token)

    return existing_tokens


def add_token_to_file(token):
    global token_cache
    if not os.path.exists(TOKEN_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Tokens"
    else:
        workbook = openpyxl.load_workbook(TOKEN_FILE)
        sheet = workbook.active

    token_cache = set(token_cache)
    token_cache.add(token)
    sheet.append([token])
    workbook.save(TOKEN_FILE)
    print(f"New token '{token}' has been added to {TOKEN_FILE}.")


def create_unique_token():
    existing_tokens = load_existing_tokens()

    while True:
        new_token = generate_unique_token()
        if new_token not in existing_tokens:
            add_token_to_file(new_token)
            return new_token
        else:
            print(f"Token {new_token} already exists, generating a new one...")


@app.route("/generatetoken", methods=["GET"])
def generate_token():
    api_key = request.args.get("api_key")
    if api_key != API_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    return f"Token: {create_unique_token()} added to the token file."


@app.route("/showtokens", methods=["GET"])
def get_tokens():
    api_key = request.args.get("api_key")
    if api_key != API_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    existing_tokens = token_cache
    return jsonify(list(existing_tokens))


@app.route("/showusers", methods=["GET"])
def get_users():
    if not os.path.exists(USER_FILE):
        return jsonify([])
    workbook = openpyxl.load_workbook(USER_FILE)
    sheet = workbook.active
    users = [
        {"username": row[0].value, "token": row[1].value}
        for row in sheet.iter_rows(min_row=2, max_col=2)
    ]
    return jsonify(users)


@app.route("/", methods=["GET"])
def hello_world():
    return "AuthBot OK!"


def load_tokens():
    global token_cache
    if not os.path.exists(TOKEN_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        workbook.save(TOKEN_FILE)
        token_cache = set()
    if not os.path.exists(USER_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Username", "Token"])
        workbook.save(USER_FILE)

    workbook = openpyxl.load_workbook(TOKEN_FILE)
    sheet = workbook.active
    token_cache = [row[0].value for row in sheet.iter_rows(min_row=1, max_col=1)]
    if token_cache[0] == None:
        token_cache = []
    # print(token_cache)
    # workbook = openpyxl.load_workbook(USER_FILE)
    # sheet = workbook.active
    # print(
    #     [(row[0].value, row[1].value) for row in sheet.iter_rows(min_row=2, max_col=2)]
    # )


def save_tokens(tokens):
    if not os.path.exists(TOKEN_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Tokens"
        workbook.save(TOKEN_FILE)
    lock_file = f"{TOKEN_FILE}.lock"

    with FileLock(lock_file):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for token in tokens:
            sheet.append([token])
        workbook.save(TOKEN_FILE)


def log_user(username, token):
    if not os.path.exists(USER_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Username", "Token"])
    else:
        workbook = openpyxl.load_workbook(USER_FILE)
        sheet = workbook.active

    sheet.append([username, token])
    workbook.save(USER_FILE)


@bot.event
async def on_ready():
    print(f"Bot is ready. Logged in as {bot.user}")
    load_tokens()


@bot.event
async def on_message(message):
    if message.author == bot.user:
        return

    if message.content.startswith("!"):
        await bot.process_commands(message)
        return

    if message.channel.name == "verify":
        global token_cache
        submitted_token = message.content.strip()

        if len(submitted_token) != 10:
            await message.channel.send(
                "❌ Invalid token. Token must be exactly 10 characters long."
            )
            return

        role_names_to_check = ["Admin", "Mod", "Bot", ROLE_NAME]

        for role_name in role_names_to_check:
            role = discord.utils.get(message.guild.roles, name=role_name)
            if role and role in message.author.roles:
                if role_name == ROLE_NAME:
                    await message.channel.send(
                        f"⚠️ {message.author.mention}, you are already a Subscriber."
                    )
                else:
                    await message.channel.send(
                        f"⚠️ {message.author.mention}, you are already in a privileged role."
                    )
                return

        if submitted_token in token_cache:
            role = discord.utils.get(message.guild.roles, name=ROLE_NAME)
            if role:
                await message.author.add_roles(role)
                await message.channel.send(
                    f"✅ {message.author.mention}, you have been verified and assigned the Subscriber role!"
                )

                token_cache = set(token_cache)
                token_cache.remove(submitted_token)
                save_tokens(token_cache)

                log_user(str(message.author), submitted_token)
            else:
                await message.channel.send(
                    f'❌ Role "{ROLE_NAME}" not found. Please contact an admin.'
                )
        else:
            await message.channel.send(
                "❌ Invalid token. Please check your token and try again."
            )

    await bot.process_commands(message)


@bot.command(name="cleanverify")
@commands.has_permissions(manage_messages=True)
async def clean_verify(ctx, limit: int):
    if ctx.channel.name != "verify":
        await ctx.send(
            "❌ This command can only be used in the #verify channel.", delete_after=5
        )
        return

    total_deleted = 0
    while limit > 0:
        delete_count = min(limit, 100)
        deleted = await ctx.channel.purge(limit=delete_count)
        total_deleted += len(deleted)
        limit -= delete_count

    await ctx.send(f"✅ Deleted {total_deleted} messages.", delete_after=5)


@clean_verify.error
async def clean_verify_error(ctx, error):
    if isinstance(error, commands.MissingPermissions):
        await ctx.send(
            "❌ You do not have permission to run this command.", delete_after=5
        )


if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())


def run_flask():
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 8080)), threaded=True)


def main():
    load_tokens()

    flask_thread = threading.Thread(target=run_flask)
    flask_thread.start()

    asyncio.run(bot.start(TOKEN))


main()
