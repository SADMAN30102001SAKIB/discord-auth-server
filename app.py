import os
import random
import string

import openpyxl
from filelock import FileLock
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

API_KEY = os.getenv("API_KEY", "TapSs@14023010.com")
TOKEN_FILE = "tokens.xlsx"
USER_FILE = "users.xlsx"
tokens_cache = set()


def generate_unique_token(length=10):
    return "".join(random.choices(string.ascii_letters + string.digits, k=length))


def load_existing_tokens():
    existing_tokens = set()

    if os.path.exists(TOKEN_FILE):
        workbook = openpyxl.load_workbook(TOKEN_FILE)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=1, max_col=1):
            token = row[0].value
            if token:
                existing_tokens.add(token)

    if os.path.exists(USER_FILE):
        workbook = openpyxl.load_workbook(USER_FILE)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, max_col=2):
            token = row[1].value
            if token:
                existing_tokens.add(token)

    return existing_tokens


def create_unique_token():
    existing_tokens = load_existing_tokens()

    while True:
        new_token = generate_unique_token()
        if new_token not in existing_tokens:
            add_token_to_file(new_token)
            return new_token
        else:
            print(f"Token {new_token} already exists, generating a new one...")


def add_token_to_file(token):
    global tokens_cache
    lock_file = f"{TOKEN_FILE}.lock"
    with FileLock(lock_file):
        if not os.path.exists(TOKEN_FILE):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Tokens"
        else:
            workbook = openpyxl.load_workbook(TOKEN_FILE)
            sheet = workbook.active

        tokens_cache = set(tokens_cache)
        tokens_cache.add(token)
        sheet.append([token])
        workbook.save(TOKEN_FILE)


@app.route("/generatetoken", methods=["GET"])
def generate_token():
    api_key = request.args.get("api_key")
    if api_key != API_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    return jsonify({"generated token": create_unique_token()})


@app.route("/showtokens", methods=["GET"])
def get_tokens():
    api_key = request.args.get("api_key")
    if api_key != API_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    existing_tokens = tokens_cache
    return jsonify(list(existing_tokens))


@app.route("/showusers", methods=["GET"])
def get_users():
    api_key = request.args.get("api_key")
    if api_key != API_KEY:
        return jsonify({"error": "Unauthorized"}), 403

    if not os.path.exists(USER_FILE):
        return jsonify([])

    workbook = openpyxl.load_workbook(USER_FILE)
    sheet = workbook.active
    users = [
        {"username": row[0].value, "token": row[1].value}
        for row in sheet.iter_rows(min_row=2, max_col=2)
    ]
    return jsonify(users)


@app.route("/downloadtokens", methods=["GET"])
def download_token():
    api_key = request.args.get("api_key")
    if api_key != API_KEY:
        return jsonify({"error": "Unauthorized"}), 403
    return send_file(TOKEN_FILE, as_attachment=True)


@app.route("/downloadusers", methods=["GET"])
def download_users():
    api_key = request.args.get("api_key")
    if api_key != API_KEY:
        return jsonify({"error": "Unauthorized"}), 403
    return send_file(USER_FILE, as_attachment=True)


@app.route("/loguser", methods=["POST"])
def log_user():
    data = request.json
    username = data.get("username")
    token = data.get("token")

    with FileLock(f"{USER_FILE}.lock"):
        if not os.path.exists(USER_FILE):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Username", "Token"])
        else:
            workbook = openpyxl.load_workbook(USER_FILE)
            sheet = workbook.active

        sheet.append([username, token])
        workbook.save(USER_FILE)
    return jsonify({"message": "User logged successfully"})


@app.route("/savetokens", methods=["POST"])
def save_tokens_api():
    tokens = request.json.get("tokens", [])

    save_tokens(tokens)
    return jsonify({"message": "Tokens saved successfully"})


def save_tokens(tokens):
    if not os.path.exists(TOKEN_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Tokens"
        workbook.save(TOKEN_FILE)

    lock_file = f"{TOKEN_FILE}.lock"
    with FileLock(lock_file):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for token in tokens:
            sheet.append([token])
        workbook.save(TOKEN_FILE)
    load_tokens()


@app.route("/", methods=["GET"])
def hello_world():
    return "AuthBot OK!"


def load_tokens():
    global tokens_cache
    if not os.path.exists(TOKEN_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        workbook.save(TOKEN_FILE)
        tokens_cache = set()
    if not os.path.exists(USER_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Username", "Token"])
        workbook.save(USER_FILE)

    workbook = openpyxl.load_workbook(TOKEN_FILE)
    sheet = workbook.active
    tokens_cache = [row[0].value for row in sheet.iter_rows(min_row=1, max_col=1)]
    if tokens_cache[0] == None:
        tokens_cache = []


if __name__ == "__main__":
    load_tokens()
    app.run(port=int(os.environ.get("PORT", 8080)))
